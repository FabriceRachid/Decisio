from pathlib import Path
import json
from io import BytesIO
from datetime import datetime

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import override_settings
from rest_framework import status
from rest_framework.test import APITestCase
from openpyxl import Workbook


TEST_MEDIA_ROOT = Path(__file__).resolve().parents[2] / 'test_media'


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT)
class IngestionAPITest(APITestCase):
    def setUp(self):
        self.password = 'testpass123'
        self.analyst = User.objects.create_user(
            username='analyst_user',
            email='analyst@example.com',
            password=self.password,
        )
        self.analyst.profile.role = 'analyst'
        self.analyst.profile.save()

        self.viewer = User.objects.create_user(
            username='viewer_user',
            email='viewer@example.com',
            password=self.password,
        )
        self.viewer.profile.role = 'viewer'
        self.viewer.profile.save()

    def _login(self, username, password):
        response = self.client.post('/api/auth/login/', {
            'username': username,
            'password': password,
        }, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        return response.data['access_token']

    def test_csv_upload_creates_data_source_and_rows(self):
        access_token = self._login(self.analyst.username, self.password)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {access_token}')

        csv_content = b"sku,amount,region\nA1,120,North\nB2,85,West\n"
        upload = SimpleUploadedFile('erp_export.csv', csv_content, content_type='text/csv')

        response = self.client.post('/api/ingestion/sources/upload/', {
            'file': upload,
            'description': 'ERP export',
            'tags': ['erp', 'finance'],
            'delimiter': ',',
            'encoding': 'utf-8',
            'has_header': 'true',
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['source_type'], 'csv')
        self.assertEqual(response.data['row_count'], 2)
        self.assertEqual(response.data['column_count'], 3)
        self.assertEqual(len(response.data['sample_rows']), 2)
        self.assertEqual(response.data['sample_rows'][0]['data']['sku'], 'A1')
        self.assertIn('schema_profile', response.data['metadata'])
        self.assertEqual(response.data['metadata']['validation_summary']['error_count'], 0)

    def test_viewer_cannot_upload_sources(self):
        access_token = self._login(self.viewer.username, self.password)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {access_token}')

        csv_content = b"sku,amount\nA1,120\n"
        upload = SimpleUploadedFile('viewer_export.csv', csv_content, content_type='text/csv')

        response = self.client.post('/api/ingestion/sources/upload/', {
            'file': upload,
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_user_only_sees_own_sources(self):
        analyst_token = self._login(self.analyst.username, self.password)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {analyst_token}')
        analyst_upload = SimpleUploadedFile('analyst.csv', b"name\nalpha\n", content_type='text/csv')
        self.client.post('/api/ingestion/sources/upload/', {'file': analyst_upload}, format='multipart')

        viewer_token = self._login(self.viewer.username, self.password)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {viewer_token}')
        list_response = self.client.get('/api/ingestion/sources/')

        self.assertEqual(list_response.status_code, status.HTTP_200_OK)
        self.assertEqual(list_response.data['count'], 0)

    def test_excel_upload_creates_data_source_and_rows(self):
        access_token = self._login(self.analyst.username, self.password)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {access_token}')

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'ERP Export'
        sheet.append(['sku', 'amount', 'region'])
        sheet.append(['A1', 120, 'North'])
        sheet.append(['B2', 85, 'West'])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        excel_file = SimpleUploadedFile(
            'erp_export.xlsx',
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response = self.client.post('/api/ingestion/sources/upload/', {
            'file': excel_file,
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['source_type'], 'excel')
        self.assertEqual(response.data['row_count'], 2)
        self.assertEqual(response.data['column_count'], 3)
        self.assertEqual(response.data['sample_rows'][1]['data']['region'], 'West')

    def test_json_upload_creates_data_source_and_rows(self):
        access_token = self._login(self.analyst.username, self.password)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {access_token}')

        payload = json.dumps([
            {'order_id': 'ERP-1', 'amount': 100.5, 'currency': 'USD'},
            {'order_id': 'ERP-2', 'amount': 75.0, 'currency': 'EUR'},
        ]).encode('utf-8')
        upload = SimpleUploadedFile('erp_export.json', payload, content_type='application/json')

        response = self.client.post('/api/ingestion/sources/upload/', {
            'file': upload,
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['source_type'], 'json')
        self.assertEqual(response.data['row_count'], 2)
        self.assertEqual(response.data['sample_rows'][0]['data']['order_id'], 'ERP-1')

    def test_preview_reports_missing_columns_and_duplicate_keys(self):
        access_token = self._login(self.analyst.username, self.password)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {access_token}')

        csv_content = b"sku,amount\nA1,120\nA1,150\n"
        upload = SimpleUploadedFile('preview.csv', csv_content, content_type='text/csv')

        response = self.client.post('/api/ingestion/sources/preview/', {
            'file': upload,
            'required_columns': ['sku', 'amount', 'region'],
            'key_columns': ['sku'],
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertFalse(response.data['can_import'])
        self.assertEqual(response.data['metadata']['duplicate_summary']['duplicate_key_count'], 2)
        self.assertTrue(any(item['code'] == 'missing_required_columns' for item in response.data['validation_errors']))
        self.assertTrue(any(item['code'] == 'duplicate_key_rows_detected' for item in response.data['validation_errors']))

    def test_preview_serializes_excel_datetime_without_http_500(self):
        access_token = self._login(self.analyst.username, self.password)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {access_token}')

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'ERP Export'
        sheet.append(['sku', 'document_date', 'amount'])
        sheet.append(['A1', datetime(2026, 4, 1, 10, 30), 120])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        excel_file = SimpleUploadedFile(
            'preview_datetime.xlsx',
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response = self.client.post('/api/ingestion/sources/preview/', {
            'file': excel_file,
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['row_count'], 1)
        self.assertIn('sample_rows', response.data)

    def test_preview_auto_detects_semicolon_csv_and_cp1252_encoding(self):
        access_token = self._login(self.analyst.username, self.password)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {access_token}')

        csv_content = "nom;ville\nCafé;Ouagadougou\n".encode('cp1252')
        upload = SimpleUploadedFile('clients.csv', csv_content, content_type='text/csv')

        response = self.client.post('/api/ingestion/sources/preview/', {
            'file': upload,
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['source_type'], 'csv')
        self.assertEqual(response.data['row_count'], 1)
        self.assertEqual(response.data['delimiter'], ';')
        self.assertEqual(response.data['encoding'], 'cp1252')
        self.assertEqual(response.data['sample_rows'][0]['data']['nom'], 'Café')

    def test_upload_detects_real_excel_even_if_extension_is_csv(self):
        access_token = self._login(self.analyst.username, self.password)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {access_token}')

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['sku', 'amount'])
        sheet.append(['A1', 120])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        disguised_excel = SimpleUploadedFile(
            'erp_disguise.csv',
            buffer.read(),
            content_type='application/octet-stream'
        )

        response = self.client.post('/api/ingestion/sources/upload/', {
            'file': disguised_excel,
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['source_type'], 'excel')
        self.assertEqual(response.data['row_count'], 1)

    def test_user_without_profile_gets_forbidden_instead_of_server_error(self):
        password = 'testpass123'
        orphan = User.objects.create_user(
            username='no_profile_user',
            email='orphan@example.com',
            password=password,
        )
        orphan.profile.delete()

        access_token = self._login(orphan.username, password)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {access_token}')

        response = self.client.get('/api/ingestion/sources/')
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_strict_validation_blocks_import_when_required_columns_missing(self):
        access_token = self._login(self.analyst.username, self.password)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {access_token}')

        csv_content = b"sku,amount\nA1,120\n"
        upload = SimpleUploadedFile('strict.csv', csv_content, content_type='text/csv')

        response = self.client.post('/api/ingestion/sources/upload/', {
            'file': upload,
            'required_columns': ['sku', 'amount', 'region'],
            'strict_validation': 'true',
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('Strict validation failed', response.data['error'])

    def test_sales_template_maps_erp_aliases_to_canonical_columns(self):
        access_token = self._login(self.analyst.username, self.password)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {access_token}')

        csv_content = (
            b"Invoice Number,Client,Posting Date,Total Amount,Devise\n"
            b"INV-1001,CUST-01,2026-03-01,450.75,USD\n"
        )
        upload = SimpleUploadedFile('sales_export.csv', csv_content, content_type='text/csv')

        response = self.client.post('/api/ingestion/sources/preview/', {
            'file': upload,
            'template_id': 'erp_sales_export',
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['can_import'])
        self.assertEqual(
            response.data['metadata']['column_mapping'],
            {
                'invoice_number': 'document_no',
                'client': 'customer_code',
                'posting_date': 'document_date',
                'total_amount': 'amount',
                'devise': 'currency',
            }
        )
        self.assertEqual(
            response.data['sample_rows'][0]['data']['document_no'],
            'INV-1001'
        )
        self.assertTrue(any(item['code'] == 'template_applied' for item in response.data['validation_errors']))

    def test_manual_column_mapping_allows_inventory_template_import(self):
        access_token = self._login(self.analyst.username, self.password)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {access_token}')

        csv_content = b"article,depot,qty\nSKU-001,WH-1,15\n"
        upload = SimpleUploadedFile('inventory.csv', csv_content, content_type='text/csv')

        response = self.client.post('/api/ingestion/sources/upload/', {
            'file': upload,
            'template_id': 'erp_inventory_export',
            'column_mapping': json.dumps({
                'article': 'sku',
                'depot': 'warehouse_code',
                'qty': 'stock_quantity',
            }),
            'strict_validation': 'true',
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['metadata']['template']['id'], 'erp_inventory_export')


class DataSourceUpdateDeleteTest(APITestCase):
    """Tests for update/delete endpoints"""
    
    def setUp(self):
        self.password = 'testpass123'
        self.analyst = User.objects.create_user(
            username='analyst_user',
            email='analyst@example.com',
            password=self.password,
        )
        self.analyst.profile.role = 'analyst'
        self.analyst.profile.save()

        self.other_analyst = User.objects.create_user(
            username='other_analyst',
            email='other@example.com',
            password=self.password,
        )
        self.other_analyst.profile.role = 'analyst'
        self.other_analyst.profile.save()

    def _login(self, username, password):
        response = self.client.post('/api/auth/login/', {
            'username': username,
            'password': password,
        }, format='json')
        return response.data['access_token']

    def _upload_source(self, user):
        """Helper to upload a test CSV"""
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {self._login(user.username, self.password)}')
        csv_content = b"sku,amount\nA1,120\nB2,85\n"
        upload = SimpleUploadedFile('test.csv', csv_content, content_type='text/csv')
        response = self.client.post('/api/ingestion/sources/upload/', {
            'file': upload,
            'description': 'Test source',
            'tags': ['test'],
        }, format='multipart')
        return response.data['id']

    def test_analyst_can_update_own_source(self):
        source_id = self._upload_source(self.analyst)

        response = self.client.put(f'/api/ingestion/sources/{source_id}/', {
            'name': 'Updated Name',
            'description': 'Updated description',
            'tags': ['updated', 'test'],
            'retention_days': 180,
        }, format='json')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['name'], 'Updated Name')
        self.assertEqual(response.data['description'], 'Updated description')
        self.assertEqual(response.data['retention_days'], 180)

    def test_analyst_cannot_update_others_source(self):
        source_id = self._upload_source(self.analyst)
        
        # Login as different analyst
        token = self._login(self.other_analyst.username, self.password)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {token}')

        response = self.client.put(f'/api/ingestion/sources/{source_id}/', {
            'name': 'Hacked',
        }, format='json')

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_analyst_can_delete_own_source_soft_delete(self):
        source_id = self._upload_source(self.analyst)

        response = self.client.delete(f'/api/ingestion/sources/{source_id}/')

        self.assertEqual(response.status_code, status.HTTP_204_NO_CONTENT)
        
        # Check source is archived, not deleted
        source = DataSource.objects.get(id=source_id)
        self.assertTrue(source.is_archived)

    def test_analyst_cannot_delete_others_source(self):
        source_id = self._upload_source(self.analyst)
        
        token = self._login(self.other_analyst.username, self.password)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {token}')

        response = self.client.delete(f'/api/ingestion/sources/{source_id}/')

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)


class DataSourceFilteringTest(APITestCase):
    """Tests for advanced filtering and search"""
    
    def setUp(self):
        self.password = 'testpass123'
        self.analyst = User.objects.create_user(
            username='analyst_user',
            email='analyst@example.com',
            password=self.password,
        )
        self.analyst.profile.role = 'analyst'
        self.analyst.profile.save()

        token = self._login(self.analyst.username, self.password)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {token}')

        # Create test sources
        self._create_source('Sales Export', 'csv', ['erp', 'sales'])
        self._create_source('Inventory Export', 'excel', ['erp', 'inventory'])
        self._create_source('Customer List', 'csv', ['master'])

    def _login(self, username, password):
        response = self.client.post('/api/auth/login/', {
            'username': username,
            'password': password,
        }, format='json')
        return response.data['access_token']

    def _create_source(self, name, source_type, tags):
        if source_type == 'csv':
            content = b"col1,col2\na,b\n"
            filename = f"{name}.csv"
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(['col1', 'col2'])
            sheet.append(['a', 'b'])
            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)
            content = buffer.read()
            filename = f"{name}.xlsx"

        upload = SimpleUploadedFile(filename, content, content_type='text/csv')
        self.client.post('/api/ingestion/sources/upload/', {
            'file': upload,
            'description': name,
            'tags': tags,
        }, format='multipart')

    def test_filter_by_source_type(self):
        response = self.client.get('/api/ingestion/sources/?source_type=csv')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['count'], 2)

    def test_filter_by_tags(self):
        response = self.client.get('/api/ingestion/sources/?tags=erp')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['count'], 2)

    def test_search_by_name(self):
        response = self.client.get('/api/ingestion/sources/?search=Sales')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertGreaterEqual(response.data['count'], 1)

    def test_pagination(self):
        response = self.client.get('/api/ingestion/sources/?page=1')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('count', response.data)
        self.assertIn('next', response.data)
        self.assertIn('previous', response.data)


class CleanupExpiredSourcesTest(TestCase):
    """Tests for cleanup command"""
    
    def setUp(self):
        self.user = User.objects.create_user(username='testuser', password='pass')

    def test_cleanup_command_removes_expired_sources(self):
        from datetime import timedelta
        from django.utils import timezone
        from django.core.management import call_command
        from django.test import override_settings
        from io import StringIO

        # Create expired source
        now = timezone.now()
        old_date = now - timedelta(days=91)
        
        source = DataSource.objects.create(
            name='Old Source',
            source_type='csv',
            file_path='test_file.csv',
            uploaded_by=self.user,
            retention_days=90,
            status='completed',
            created_at=old_date,
        )

        # Run cleanup
        out = StringIO()
        call_command('cleanup_expired_sources', '--force', stdout=out)
        
        # Check source is deleted
        self.assertFalse(DataSource.objects.filter(id=source.id).exists())

    def test_cleanup_dry_run_doesnt_delete(self):
        from datetime import timedelta
        from django.utils import timezone
        from django.core.management import call_command
        from io import StringIO

        now = timezone.now()
        old_date = now - timedelta(days=91)
        
        source = DataSource.objects.create(
            name='Old Source',
            source_type='csv',
            file_path='test_file.csv',
            uploaded_by=self.user,
            retention_days=90,
            status='completed',
            created_at=old_date,
        )

        # Run cleanup with --dry-run
        out = StringIO()
        call_command('cleanup_expired_sources', '--dry-run', '--force', stdout=out)
        
        # Check source still exists
        self.assertTrue(DataSource.objects.filter(id=source.id).exists())
        self.assertEqual(response.data['sample_rows'][0]['data']['sku'], 'SKU-001')
        self.assertEqual(response.data['sample_rows'][0]['data']['warehouse_code'], 'WH-1')
        self.assertEqual(response.data['sample_rows'][0]['data']['stock_quantity'], 15)

    def test_template_catalog_endpoint_lists_available_presets(self):
        access_token = self._login(self.analyst.username, self.password)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {access_token}')

        response = self.client.get('/api/ingestion/templates/')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        template_ids = {item['id'] for item in response.data['results']}
        self.assertIn('erp_sales_export', template_ids)
        self.assertIn('erp_inventory_export', template_ids)
        self.assertIn('erp_customer_master', template_ids)
        self.assertIn('erp_supplier_master', template_ids)
        self.assertIn('erp_gl_entries', template_ids)

    def test_sales_template_business_rules_flag_invalid_currency_and_amount(self):
        access_token = self._login(self.analyst.username, self.password)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {access_token}')

        csv_content = (
            b"Invoice Number,Client,Posting Date,Total Amount,Devise\n"
            b"INV-1001,CUST-01,2026-03-01,-10,XYZ\n"
        )
        upload = SimpleUploadedFile('sales_rules.csv', csv_content, content_type='text/csv')

        response = self.client.post('/api/ingestion/sources/preview/', {
            'file': upload,
            'template_id': 'erp_sales_export',
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        row_messages = response.data['sample_rows'][0]['messages']
        self.assertTrue(any(item['code'] == 'positive_number' for item in row_messages))
        self.assertTrue(any(item['code'] == 'allowed_values' for item in row_messages))
        self.assertEqual(response.data['sample_rows'][0]['status'], 'warning')

    def test_inventory_template_business_rules_flag_negative_stock(self):
        access_token = self._login(self.analyst.username, self.password)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {access_token}')

        csv_content = b"article,depot,qty\nSKU-001,WH-1,-3\n"
        upload = SimpleUploadedFile('inventory_rules.csv', csv_content, content_type='text/csv')

        response = self.client.post('/api/ingestion/sources/preview/', {
            'file': upload,
            'template_id': 'erp_inventory_export',
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        row_messages = response.data['sample_rows'][0]['messages']
        self.assertTrue(any(item['code'] == 'non_negative_number' for item in row_messages))
        self.assertEqual(response.data['sample_rows'][0]['status'], 'warning')
