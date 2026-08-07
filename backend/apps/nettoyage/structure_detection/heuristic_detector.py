"""
Heuristic structure detection for messy Excel/CSV files.
Reads files via openpyxl (for merged cells and formatting) and pandas.
Returns a structural fingerprint + confidence score.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd


@dataclass
class SubTable:
    start_row: int
    end_row: int
    start_col: int
    end_col: int
    header_row: int | None = None
    column_names: list[str] = field(default_factory=list)
    confidence: float = 0.0


@dataclass
class StructuralFingerprint:
    total_rows: int = 0
    total_cols: int = 0
    merged_cells: list[dict[str, int]] = field(default_factory=list)
    blank_rows: list[int] = field(default_factory=list)
    blank_cols: list[int] = field(default_factory=list)
    header_candidates: list[dict[str, Any]] = field(default_factory=list)
    subtables: list[SubTable] = field(default_factory=list)
    column_types: dict[str, str] = field(default_factory=dict)
    confidence: float = 0.0
    issues: list[str] = field(default_factory=list)
    sparse_pivot_candidates: list[dict[str, Any]] = field(default_factory=list)
    hierarchical_headers: list[dict[str, Any]] = field(default_factory=list)
    hierarchical_crosstab: dict[str, Any] | None = None
    force_llm: bool = False

    def to_dict(self) -> dict[str, Any]:
        return {
            'total_rows': self.total_rows,
            'total_cols': self.total_cols,
            'merged_cells': self.merged_cells,
            'blank_rows': self.blank_rows,
            'blank_cols': self.blank_cols,
            'header_candidates': self.header_candidates,
            'subtables': [
                {
                    'start_row': st.start_row,
                    'end_row': st.end_row,
                    'start_col': st.start_col,
                    'end_col': st.end_col,
                    'header_row': st.header_row,
                    'column_names': st.column_names,
                    'confidence': st.confidence,
                }
                for st in self.subtables
            ],
            'column_types': self.column_types,
            'confidence': self.confidence,
            'issues': self.issues,
            'sparse_pivot_candidates': self.sparse_pivot_candidates,
            'hierarchical_headers': self.hierarchical_headers,
            'hierarchical_crosstab': self.hierarchical_crosstab,
            'force_llm': self.force_llm,
            '_sample_data': getattr(self, '_sample_data', None),
        }


class HeuristicDetector:
    """
    Detects structure of Excel/CSV files using heuristic rules.
    No LLM calls — fast, deterministic analysis.
    """

    BLANK_THRESHOLD = 0.6
    HEADER_MIN_NON_EMPTY = 0.5
    SUBTABLE_MIN_ROWS = 3
    SUBTABLE_MIN_COLS = 2

    def detect_from_file(self, file_path: str, sheet_name: str | None = None) -> StructuralFingerprint:
        path = Path(file_path)
        suffix = path.suffix.lower()

        if suffix in ('.xlsx', '.xls'):
            return self._detect_excel(path, sheet_name)
        elif suffix == '.csv':
            return self._detect_csv(path)
        else:
            raise ValueError(f"Unsupported file type: {suffix}")

    def detect_from_dataframe(self, df: pd.DataFrame, sheet_name: str = '') -> StructuralFingerprint:
        fp = StructuralFingerprint()
        fp.total_rows = len(df)
        fp.total_cols = len(df.columns)

        fp.column_types = self._infer_column_types(df)
        fp.header_candidates = self._find_header_candidates_df(df)
        fp.subtables = self._detect_subtables_from_blanks(df)
        fp.confidence = self._compute_confidence(fp)
        fp.issues = self._detect_issues(df, fp)

        import numpy as np
        header_row = tuple(str(c) for c in df.columns)
        all_rows = [header_row] + [tuple(None if pd.isna(v) else v for v in df.iloc[i]) for i in range(len(df))]
        fp.sparse_pivot_candidates = self.detect_sparse_pivot_pattern(all_rows, 0)
        fp.hierarchical_headers = self.detect_hierarchical_headers(all_rows, [])
        fp.hierarchical_crosstab = self.detect_hierarchical_crosstab(all_rows)
        if fp.sparse_pivot_candidates or fp.hierarchical_crosstab:
            fp.force_llm = True

        return fp

    def _detect_excel(self, path: Path, sheet_name: str | None = None) -> StructuralFingerprint:
        wb = openpyxl.load_workbook(path, data_only=False)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        fp = StructuralFingerprint()
        fp.total_rows = ws.max_row or 0
        fp.total_cols = ws.max_column or 0

        try:
            fp.merged_cells = [
                {
                    'start_row': m.min_row,
                    'start_col': m.min_col,
                    'end_row': m.max_row,
                    'end_col': m.max_col,
                }
                for m in ws.merged_cells.ranges
            ]
        except (AttributeError, TypeError):
            fp.merged_cells = []

        all_rows = list(ws.iter_rows(min_row=1, max_row=fp.total_rows, values_only=True))

        for m in ws.merged_cells.ranges:
            top_val = all_rows[m.min_row - 1][m.min_col - 1] if m.min_row - 1 < len(all_rows) and m.min_col - 1 < len(all_rows[m.min_row - 1]) else None
            for r in range(m.min_row - 1, min(m.max_row, len(all_rows))):
                row = all_rows[r]
                for c in range(m.min_col - 1, min(m.max_col, len(row))):
                    if (r, c) != (m.min_row - 1, m.min_col - 1):
                        all_rows[r] = row[:c] + (top_val,) + row[c + 1:]
                        row = all_rows[r]

        fp.blank_rows = self._find_blank_rows(all_rows)
        fp.blank_cols = self._find_blank_cols(all_rows, fp.total_cols)

        fp.header_candidates = self._find_header_candidates(all_rows, fp.blank_rows)
        fp.subtables = self._detect_subtables(all_rows, fp.blank_rows, fp.blank_cols, fp.merged_cells)
        fp.column_types = self._infer_types_from_values(all_rows, fp.header_candidates)
        fp.confidence = self._compute_confidence(fp)
        fp.issues = self._detect_issues_from_raw(all_rows, fp)

        best_header = fp.header_candidates[0]['row_index'] if fp.header_candidates else 0
        fp.sparse_pivot_candidates = self.detect_sparse_pivot_pattern(all_rows, best_header)
        fp.hierarchical_headers = self.detect_hierarchical_headers(all_rows, fp.blank_rows)
        fp.hierarchical_crosstab = self.detect_hierarchical_crosstab(all_rows)
        if fp.sparse_pivot_candidates or fp.hierarchical_crosstab:
            fp.force_llm = True

        def _safe(val):
            if val is None:
                return None
            if hasattr(val, 'isoformat'):
                return val.isoformat()
            if isinstance(val, (set, tuple)):
                return list(val)
            return val

        fp._sample_data = {
            'header_row': [_safe(c) for c in all_rows[best_header]] if best_header < len(all_rows) else [],
            'first_rows': [[_safe(c) for c in r] for r in all_rows[:5] if r != all_rows[best_header]],
            'last_rows': [[_safe(c) for c in r] for r in all_rows[-3:]],
            'raw_header_rows': [[_safe(c) for c in r] for r in all_rows[:6]],
        }

        wb.close()
        return fp

    def _detect_csv(self, path: Path) -> StructuralFingerprint:
        df = pd.read_csv(path, header=None, dtype=str, keep_default_na=False)
        return self.detect_from_dataframe(df)

    def _find_blank_rows(self, all_rows: list[tuple]) -> list[int]:
        blank = []
        for i, row in enumerate(all_rows):
            non_empty = sum(1 for cell in row if cell is not None and str(cell).strip())
            ratio = non_empty / len(row) if row else 0
            if ratio < self.BLANK_THRESHOLD:
                blank.append(i)
        return blank

    def _find_blank_cols(self, all_rows: list[tuple], total_cols: int) -> list[int]:
        if not all_rows:
            return []
        non_blank_rows = [
            row for row in all_rows
            if sum(1 for cell in row if cell is not None and str(cell).strip()) / len(row) >= self.BLANK_THRESHOLD
        ]
        if not non_blank_rows:
            return []
        blank = []
        for col_idx in range(total_cols):
            non_empty = sum(
                1 for row in non_blank_rows
                if col_idx < len(row) and row[col_idx] is not None and str(row[col_idx]).strip()
            )
            ratio = non_empty / len(non_blank_rows)
            if ratio < self.BLANK_THRESHOLD:
                blank.append(col_idx)
        return blank

    def _find_header_candidates(
        self, all_rows: list[tuple], blank_rows: list[int]
    ) -> list[dict[str, Any]]:
        candidates = []
        for i, row in enumerate(all_rows):
            if i in blank_rows:
                continue
            non_empty = sum(1 for cell in row if cell is not None and str(cell).strip())
            ratio = non_empty / len(row) if row else 0
            if ratio >= self.HEADER_MIN_NON_EMPTY:
                has_text = sum(
                    1 for cell in row
                    if cell is not None and isinstance(cell, str) and not cell.replace('.', '').replace('-', '').isdigit()
                )
                candidates.append({
                    'row_index': i,
                    'fill_ratio': round(ratio, 3),
                    'text_count': has_text,
                    'score': round(ratio * 0.6 + (has_text / len(row) if row else 0) * 0.4, 3),
                })
        candidates.sort(key=lambda c: c['score'], reverse=True)
        return candidates[:5]

    def _find_header_candidates_df(self, df: pd.DataFrame) -> list[dict[str, Any]]:
        candidates = []
        for i in range(min(20, len(df))):
            row = df.iloc[i]
            non_empty = row.count()
            ratio = non_empty / len(row) if len(row) > 0 else 0
            if ratio >= self.HEADER_MIN_NON_EMPTY:
                text_count = sum(
                    1 for val in row
                    if isinstance(val, str) and not val.replace('.', '').replace('-', '').isdigit()
                )
                candidates.append({
                    'row_index': i,
                    'fill_ratio': round(ratio, 3),
                    'text_count': text_count,
                    'score': round(ratio * 0.6 + (text_count / len(row) if len(row) > 0 else 0) * 0.4, 3),
                })
        candidates.sort(key=lambda c: c['score'], reverse=True)
        return candidates[:5]

    def _detect_subtables(
        self,
        all_rows: list[tuple],
        blank_rows: list[int],
        blank_cols: list[int],
        merged_cells: list[dict],
    ) -> list[SubTable]:
        if not blank_rows and not merged_cells:
            if all_rows:
                return [SubTable(
                    start_row=0, end_row=len(all_rows) - 1,
                    start_col=0, end_col=(len(all_rows[0]) if all_rows else 1) - 1,
                    header_row=0, confidence=0.9,
                )]
            return []

        segments = self._split_by_blank_rows(all_rows, blank_rows)
        subtables = []
        for seg_start, seg_end in segments:
            if (seg_end - seg_start + 1) < self.SUBTABLE_MIN_ROWS:
                continue
            cols = self._find_segment_cols(all_rows, seg_start, seg_end, blank_cols)
            if len(cols) < self.SUBTABLE_MIN_COLS:
                continue
            subtables.append(SubTable(
                start_row=seg_start,
                end_row=seg_end,
                start_col=cols[0],
                end_col=cols[-1],
                confidence=0.7,
            ))

        if not subtables and all_rows:
            cols = self._find_segment_cols(all_rows, 0, len(all_rows) - 1, blank_cols)
            if len(cols) >= self.SUBTABLE_MIN_COLS:
                subtables.append(SubTable(
                    start_row=0, end_row=len(all_rows) - 1,
                    start_col=cols[0], end_col=cols[-1],
                    confidence=0.5,
                ))

        return subtables

    def _detect_subtables_from_blanks(self, df: pd.DataFrame) -> list[SubTable]:
        blank_rows = []
        for i in range(len(df)):
            non_empty = df.iloc[i].count()
            ratio = non_empty / len(df.columns) if len(df.columns) > 0 else 0
            if ratio < self.BLANK_THRESHOLD:
                blank_rows.append(i)

        blank_cols = []
        for j in range(len(df.columns)):
            non_empty = df.iloc[:, j].count()
            ratio = non_empty / len(df)
            if ratio < self.BLANK_THRESHOLD:
                blank_cols.append(j)

        segments = []
        prev_blank = -1
        for br in blank_rows:
            if br > prev_blank + 1:
                segments.append((prev_blank + 1, br - 1))
            prev_blank = br
        if prev_blank < len(df) - 1:
            segments.append((prev_blank + 1, len(df) - 1))

        subtables = []
        for seg_start, seg_end in segments:
            if (seg_end - seg_start + 1) < self.SUBTABLE_MIN_ROWS:
                continue
            usable_cols = [j for j in range(len(df.columns)) if j not in blank_cols]
            if len(usable_cols) < self.SUBTABLE_MIN_COLS:
                continue
            subtables.append(SubTable(
                start_row=seg_start, end_row=seg_end,
                start_col=usable_cols[0], end_col=usable_cols[-1],
                confidence=0.7,
            ))

        if not subtables and len(df) > 0:
            usable_cols = [j for j in range(len(df.columns)) if j not in blank_cols]
            if len(usable_cols) >= self.SUBTABLE_MIN_COLS:
                subtables.append(SubTable(
                    start_row=0, end_row=len(df) - 1,
                    start_col=usable_cols[0], end_col=usable_cols[-1],
                    confidence=0.5,
                ))

        return subtables

    def _split_by_blank_rows(self, all_rows: list[tuple], blank_rows: list[int]) -> list[tuple[int, int]]:
        segments = []
        prev_blank = -1
        for br in sorted(blank_rows):
            if br > prev_blank + 1:
                segments.append((prev_blank + 1, br - 1))
            prev_blank = br
        if prev_blank < len(all_rows) - 1:
            segments.append((prev_blank + 1, len(all_rows) - 1))
        return segments

    def _find_segment_cols(
        self, all_rows: list[tuple], start: int, end: int, global_blank_cols: list[int]
    ) -> list[int]:
        cols = []
        for j in range(len(all_rows[0]) if all_rows else 0):
            if j in global_blank_cols:
                continue
            non_empty = sum(
                1 for i in range(start, min(end + 1, len(all_rows)))
                if j < len(all_rows[i]) and all_rows[i][j] is not None and str(all_rows[i][j]).strip()
            )
            total = end - start + 1
            if non_empty / total >= 0.3:
                cols.append(j)
        return cols

    def _infer_column_types(self, df: pd.DataFrame) -> dict[str, str]:
        types = {}
        for col in df.columns:
            sample = df[col].dropna().head(50)
            if sample.empty:
                types[str(col)] = 'empty'
                continue
            types[str(col)] = self._classify_column(sample.tolist())
        return types

    def _infer_types_from_values(
        self, all_rows: list[tuple], header_candidates: list[dict]
    ) -> dict[str, str]:
        if not header_candidates or not all_rows:
            return {}

        header_idx = header_candidates[0]['row_index']
        headers = [str(c) if c is not None else f'col_{i}' for i, c in enumerate(all_rows[header_idx])]
        data_rows = all_rows[header_idx + 1:]

        types = {}
        for j, h in enumerate(headers):
            col_values = [str(row[j]) for row in data_rows if j < len(row) and row[j] is not None]
            if col_values:
                types[h] = self._classify_column(col_values)
            else:
                types[h] = 'empty'
        return types

    def _classify_column(self, values: list) -> str:
        if not values:
            return 'empty'

        str_vals = [str(v).strip() for v in values if str(v).strip()]
        if not str_vals:
            return 'empty'

        numeric_count = sum(1 for v in str_vals if self._is_numeric(v))
        date_count = sum(1 for v in str_vals if self._is_date_like(v))
        total = len(str_vals)

        if numeric_count / total > 0.8:
            return 'numeric'
        if date_count / total > 0.8:
            return 'date'
        return 'text'

    def _is_numeric(self, val: str) -> bool:
        cleaned = val.replace(',', '').replace('.', '').replace('-', '').replace('+', '').strip()
        if not cleaned:
            return False
        return cleaned.isdigit()

    def _is_date_like(self, val: str) -> bool:
        date_patterns = [
            r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}',
            r'\d{4}[/-]\d{1,2}[/-]\d{1,2}',
            r'\d{1,2}\s+\w+\s+\d{4}',
        ]
        return any(re.match(p, val) for p in date_patterns)

    def _compute_confidence(self, fp: StructuralFingerprint) -> float:
        score = 0.5

        if fp.header_candidates:
            best = fp.header_candidates[0]
            if best['score'] > 0.7:
                score += 0.2
            elif best['score'] > 0.5:
                score += 0.1
        else:
            score -= 0.15

        if len(fp.subtables) == 1:
            score += 0.15
        elif len(fp.subtables) > 1:
            score -= 0.1 * min(len(fp.subtables) - 1, 4)

        if not fp.merged_cells:
            score += 0.1
        else:
            score -= 0.05 * min(len(fp.merged_cells), 4)

        if len(fp.blank_rows) < 3:
            score += 0.05
        elif len(fp.blank_rows) > 5:
            score -= 0.1

        empty_cols = sum(1 for t in fp.column_types.values() if t == 'empty')
        if empty_cols > 2:
            score -= 0.05

        return max(min(round(score, 4), 1.0), 0.0)

    def _detect_issues(self, df: pd.DataFrame, fp: StructuralFingerprint) -> list[str]:
        issues = []
        if len(fp.merged_cells) > 0:
            issues.append(f'{len(fp.merged_cells)} cellule(s) fusionnee(s) detectee(s)')
        if len(fp.blank_rows) > 5:
            issues.append(f'{len(fp.blank_rows)} lignes vides detectees (separateurs potentiels)')
        if len(fp.subtables) > 1:
            issues.append(f'{len(fp.subtables)} sous-tableaux detectes dans la meme feuille')
        if not fp.header_candidates:
            issues.append('Aucune ligne d\'en-tete candidate detectee')
        empty_cols = sum(1 for t in fp.column_types.values() if t == 'empty')
        if empty_cols > 0:
            issues.append(f'{empty_cols} colonne(s) vide(s) detectee(s)')
        return issues

    def detect_sparse_pivot_pattern(
        self,
        all_rows: list[tuple],
        header_row_idx: int | None = None,
    ) -> list[dict[str, Any]]:
        if not all_rows or len(all_rows) < 3:
            return []
        if header_row_idx is None:
            header_row_idx = 0
        headers = [str(c) if c is not None else '' for c in all_rows[header_row_idx]]
        data_rows = all_rows[header_row_idx + 1:]
        if not data_rows:
            return []
        numeric_cols = []
        for j in range(len(headers)):
            numeric_count = 0
            total = 0
            for row in data_rows:
                if j < len(row) and row[j] is not None and str(row[j]).strip():
                    total += 1
                    val = str(row[j]).replace(',', '').replace('.', '').replace('-', '').strip()
                    if val.isdigit():
                        numeric_count += 1
            if total > 0 and numeric_count / total > 0.7:
                numeric_cols.append(j)
        if len(numeric_cols) < 3:
            return []
        always_filled_cols = []
        for j in range(len(headers)):
            filled_count = sum(
                1 for row in data_rows
                if j < len(row) and row[j] is not None and str(row[j]).strip()
            )
            if filled_count / len(data_rows) >= 0.95:
                always_filled_cols.append(j)
        candidates = []
        window_size = min(len(numeric_cols), 8)
        for w in range(3, window_size + 1):
            for start_idx in range(len(numeric_cols) - w + 1):
                group = numeric_cols[start_idx:start_idx + w]
                single_value_rows = 0
                for row in data_rows:
                    non_empty = sum(
                        1 for ci in group
                        if ci < len(row) and row[ci] is not None and str(row[ci]).strip()
                    )
                    if non_empty == 1:
                        single_value_rows += 1
                ratio = single_value_rows / len(data_rows) if data_rows else 0
                if ratio >= 0.9:
                    dim_cols = [ci for ci in group if ci in always_filled_cols]
                    value_cols = [ci for ci in group if ci not in always_filled_cols]
                    if len(value_cols) >= 2:
                        dim_cols = []
                        for j in range(len(headers)):
                            if j in group:
                                continue
                            filled_count = sum(
                                1 for row in data_rows
                                if j < len(row) and row[j] is not None and str(row[j]).strip()
                            )
                            if filled_count / len(data_rows) >= 0.95:
                                min_col = min(group)
                                max_col = max(group)
                                if j == min_col - 1 or j == max_col + 1:
                                    dim_cols.append(j)
                        sparsity_score = 1.0 - (sum(
                            sum(
                                1 for ci in group
                                if ci < len(row) and row[ci] is not None and str(row[ci]).strip()
                            )
                            for row in data_rows
                        ) / (len(data_rows) * len(group)))
                        candidates.append({
                            'group_cols': group,
                            'group_headers': [headers[c] for c in group],
                            'dimension_cols': dim_cols,
                            'dimension_headers': [headers[c] for c in dim_cols],
                            'value_cols': value_cols,
                            'value_headers': [headers[c] for c in value_cols],
                            'single_value_ratio': round(ratio, 3),
                            'sparsity_score': round(sparsity_score, 3),
                            'data_rows': len(data_rows),
                        })
        seen = set()
        unique = []
        for c in candidates:
            key = tuple(c['value_cols'])
            if key not in seen:
                seen.add(key)
                unique.append(c)
        unique.sort(key=lambda x: x['single_value_ratio'], reverse=True)
        return unique[:3]

    def _compute_group_sparsity(self, data_rows: list[tuple], col_indices: list[int]) -> float:
        if not data_rows or not col_indices:
            return 0.0
        total_cells = len(data_rows) * len(col_indices)
        if total_cells == 0:
            return 0.0
        non_empty_count = 0
        for row in data_rows:
            for ci in col_indices:
                if ci < len(row) and row[ci] is not None and str(row[ci]).strip():
                    non_empty_count += 1
        if non_empty_count == 0:
            return 1.0
        fill_ratio = non_empty_count / total_cells
        sparsity = 1.0 - fill_ratio
        return sparsity

    def detect_hierarchical_headers(
        self, all_rows: list[tuple], blank_rows: list[int], pivot_group: dict | None = None
    ) -> list[dict[str, Any]]:
        if not all_rows or len(all_rows) < 2:
            return []
        hierarchies = []
        max_header_rows = min(5, len(all_rows))
        for h in range(1, max_header_rows):
            row = all_rows[h]
            if h in blank_rows:
                continue
            non_empty = sum(1 for c in row if c is not None and str(c).strip())
            if non_empty < 2:
                continue
            merged_under = False
            for mc in self._merged_ranges(all_rows):
                if mc['start_row'] <= h <= mc['end_row'] and mc['end_row'] > mc['start_row']:
                    merged_under = True
                    break
            if not merged_under and h > 0:
                prev_row = all_rows[h - 1]
                prev_non_empty = sum(1 for c in prev_row if c is not None and str(c).strip())
                if prev_non_empty > non_empty:
                    continue
            hierarchies.append({
                'row_index': h,
                'values': [str(c) if c is not None else '' for c in row],
                'non_empty_count': non_empty,
            })
        return hierarchies

    def _merged_ranges(self, all_rows: list[tuple]) -> list[dict[str, int]]:
        ranges = []
        for i, row in enumerate(all_rows):
            prev = None
            for j, cell in enumerate(row):
                if cell is None or (isinstance(cell, str) and not cell.strip()):
                    if prev is not None:
                        ranges.append({'start_row': i, 'start_col': prev, 'end_row': i, 'end_col': j - 1})
                        prev = None
                else:
                    prev = j
            if prev is not None:
                ranges.append({'start_row': i, 'start_col': prev, 'end_row': i, 'end_col': len(row) - 1})
        return ranges

    @staticmethod
    def _is_numeric_value(val: str) -> bool:
        if not val:
            return False
        cleaned = val.replace(',', '').replace('.', '').replace('-', '').replace(' ', '')
        if cleaned.isdigit():
            return True
        try:
            float(val.replace(',', '.'))
            return True
        except (ValueError, TypeError):
            return False

    def detect_hierarchical_crosstab(
        self, all_rows: list[tuple]
    ) -> dict[str, Any] | None:
        if not all_rows or len(all_rows) < 4:
            return None

        header_rows = []
        for i in range(min(6, len(all_rows))):
            row = all_rows[i]
            non_empty = [j for j, c in enumerate(row) if c is not None and str(c).strip()]
            if not non_empty:
                break
            is_text_row = all(
                not self._is_numeric_value(str(row[j]).strip())
                for j in non_empty
            )
            if is_text_row and len(non_empty) >= 2:
                header_rows.append((i, non_empty))
            elif header_rows:
                break

        if len(header_rows) < 2:
            return None

        first_data_row = header_rows[-1][0] + 1
        data_rows = all_rows[first_data_row:]
        if len(data_rows) < 5:
            return None

        fills_per_row = []
        for row in data_rows:
            ne = sum(1 for c in row if c is not None and str(c).strip())
            fills_per_row.append(ne)

        avg_fill = sum(fills_per_row) / len(fills_per_row) if fills_per_row else 0
        total_cols = len(all_rows[0]) if all_rows else 0
        if total_cols < 4:
            return None
        sparsity_ratio = avg_fill / total_cols
        if sparsity_ratio > 0.4 or avg_fill > 5:
            return None

        col_paths = {j: [] for j in range(total_cols)}
        for hi, (row_idx, non_empty_cols) in enumerate(header_rows):
            row = all_rows[row_idx]
            prev_val = None
            for j in range(total_cols):
                if j in non_empty_cols:
                    val = str(row[j]).strip() if row[j] is not None else ''
                    if val:
                        col_paths[j].append(val)
                        prev_val = val
                    else:
                        col_paths[j].append(prev_val or '')
                else:
                    col_paths[j].append(prev_val or '')

        key_cols = []
        for j in range(total_cols):
            if j < 2:
                filled = sum(
                    1 for r in data_rows
                    if j < len(r) and r[j] is not None and str(r[j]).strip()
                )
                if filled / len(data_rows) > 0.9:
                    key_cols.append(j)

        value_cols = []
        for j in range(total_cols):
            if j in key_cols:
                continue
            filled = sum(
                1 for r in data_rows
                if j < len(r) and r[j] is not None and str(r[j]).strip()
            )
            if 0 < filled / len(data_rows) < 0.5:
                value_cols.append(j)

        if len(value_cols) < 2:
            return None

        unpivot_map = []
        for vc in value_cols:
            path = [p for p in col_paths[vc] if p]
            unpivot_map.append({'source_col': vc, 'path': path})

        dim_headers = []
        for hi, (row_idx, non_empty_cols) in enumerate(header_rows):
            row = all_rows[row_idx]
            label = None
            for j in non_empty_cols:
                val = str(row[j]).strip() if row[j] is not None else ''
                if val and not self._is_numeric_value(val):
                    label = val
                    break
            dim_headers.append(label or f'Niveau_{hi+1}')

        return {
            'type': 'hierarchical_crosstab',
            'header_rows': [hr[0] for hr in header_rows],
            'key_cols': key_cols,
            'value_cols': value_cols,
            'unpivot_map': unpivot_map,
            'dim_headers': dim_headers,
            'data_row_count': len(data_rows),
            'sparsity_ratio': round(sparsity_ratio, 3),
        }

    def _detect_issues_from_raw(self, all_rows: list[tuple], fp: StructuralFingerprint) -> list[str]:
        issues = []
        if fp.merged_cells:
            issues.append(f'{len(fp.merged_cells)} cellule(s) fusionnee(s) detectee(s)')
        if len(fp.blank_rows) > 5:
            issues.append(f'{len(fp.blank_rows)} lignes vides detectees (separateurs potentiels)')
        if len(fp.subtables) > 1:
            issues.append(f'{len(fp.subtables)} sous-tableaux detectes dans la meme feuille')
        if not fp.header_candidates:
            issues.append('Aucune ligne d\'en-tete candidate detectee')
        if fp.sparse_pivot_candidates:
            issues.append(
                f'{len(fp.sparse_pivot_candidates)} schema(s) pivot epars detecte(s) '
                '(necessite transformation unpivot)'
            )
        if fp.hierarchical_crosstab:
            hc = fp.hierarchical_crosstab
            issues.append(
                f'Crosstab hiérarchique detecte ({hc["data_row_count"]} lignes, '
                f'{len(hc["value_cols"])} colonnes de valeurs, sparsite {hc["sparsity_ratio"]:.0%})'
            )
        return issues
