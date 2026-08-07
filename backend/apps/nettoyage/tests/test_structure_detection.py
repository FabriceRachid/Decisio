"""
Tests for intelligent structural reconstruction (structure_detection).
"""
import json
import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import numpy as np
import pandas as pd
import pytest
from openpyxl import Workbook

from apps.nettoyage.structure_detection.heuristic_detector import HeuristicDetector, StructuralFingerprint
from apps.nettoyage.structure_detection.validation_gates import ValidationGates, ValidationReport
from apps.nettoyage.structure_detection.llm_service import LLMReconstructionService
from apps.nettoyage.structure_detection.correction_memory import CorrectionMemory
from apps.nettoyage.structure_detection.pivot_transformer import PivotTransformer
from apps.nettoyage.structure_detection.cell_transformer import (
    extract_labeled_fields,
    fix_ambiguous_numeric_chars,
    split_value_unit,
    explode_delimited_lists,
    CellTransformerEngine,
)
from apps.nettoyage.structure_detection.pivot_transformer import PivotTransformer


# ── Heuristic Detector Tests ──

class TestHeuristicDetector:
    def setup_method(self):
        self.detector = HeuristicDetector()

    def _create_clean_excel(self, tmp_path, data, sheet_name="Sheet1"):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        for row in data:
            ws.append(row)
        path = tmp_path / "test_clean.xlsx"
        wb.save(path)
        return str(path)

    def _create_messy_excel(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.merge_cells("A1:C1")
        ws["A1"] = "Titre du rapport"
        ws.append(["", "", ""])
        ws.append(["Nom", "Prenom", "Montant"])
        ws.append(["Dupont", "Jean", 100])
        ws.append(["Martin", "Marie", 200])
        ws.append(["", "", ""])
        ws.append(["Resume", "", ""])
        ws.append(["Total", "", 300])
        path = tmp_path / "test_messy.xlsx"
        wb.save(path)
        return str(path)

    def _create_multi_table_excel(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Client", "Ville", "Montant"])
        ws.append(["A", "Paris", 100])
        ws.append(["B", "Lyon", 200])
        ws.append(["", "", ""])
        ws.append(["", "", ""])
        ws.append(["Produit", "Prix", "Stock"])
        ws.append(["X", 50, 10])
        ws.append(["Y", 30, 20])
        path = tmp_path / "test_multi.xlsx"
        wb.save(path)
        return str(path)

    def test_detect_clean_excel(self, tmp_path):
        data = [
            ["Nom", "Prenom", "Age"],
            ["Alice", "Bob", 25],
            ["Charlie", "Delta", 30],
            ["Eve", "Frank", 35],
        ]
        path = self._create_clean_excel(tmp_path, data)
        fp = self.detector.detect_from_file(path)

        assert fp.total_rows == 4
        assert fp.total_cols == 3
        assert fp.confidence >= 0.5
        assert len(fp.merged_cells) == 0

    def test_detect_merged_cells(self, tmp_path):
        path = self._create_messy_excel(tmp_path)
        fp = self.detector.detect_from_file(path)

        assert len(fp.merged_cells) > 0
        assert fp.merged_cells[0]['start_row'] == 1
        assert fp.merged_cells[0]['start_col'] == 1

    def test_detect_blank_rows(self, tmp_path):
        path = self._create_messy_excel(tmp_path)
        fp = self.detector.detect_from_file(path)

        assert len(fp.blank_rows) > 0

    def test_detect_multiple_subtables(self, tmp_path):
        path = self._create_multi_table_excel(tmp_path)
        fp = self.detector.detect_from_file(path)

        assert len(fp.subtables) >= 1

    def test_detect_from_dataframe(self):
        df = pd.DataFrame({
            "Nom": ["Alice", "Bob", "Charlie"],
            "Age": [25, 30, 35],
            "Ville": ["Paris", "Lyon", "Marseille"],
        })
        fp = self.detector.detect_from_dataframe(df)

        assert fp.total_rows == 3
        assert fp.total_cols == 3
        assert fp.confidence > 0

    def test_header_candidates(self, tmp_path):
        data = [
            ["Nom", "Prenom", "Montant"],
            ["Alice", "Bob", 100],
            ["Charlie", "Delta", 200],
        ]
        path = self._create_clean_excel(tmp_path, data)
        fp = self.detector.detect_from_file(path)

        assert len(fp.header_candidates) > 0
        assert fp.header_candidates[0]['row_index'] == 0

    def test_confidence_score_range(self, tmp_path):
        data = [
            ["A", "B", "C"],
            ["1", "2", "3"],
            ["4", "5", "6"],
        ]
        path = self._create_clean_excel(tmp_path, data)
        fp = self.detector.detect_from_file(path)

        assert 0 <= fp.confidence <= 1

    def test_to_dict(self, tmp_path):
        data = [["X", "Y"], ["1", "2"]]
        path = self._create_clean_excel(tmp_path, data)
        fp = self.detector.detect_from_file(path)
        d = fp.to_dict()

        assert isinstance(d, dict)
        assert 'total_rows' in d
        assert 'confidence' in d
        assert 'subtables' in d


# ── Validation Gates Tests ──

class TestValidationGates:
    def setup_method(self):
        self.gates = ValidationGates()

    def test_high_confidence_passes(self):
        plan = {'confidence': 0.9, 'subtables': [{'start_row': 0, 'end_row': 5, 'start_col': 0, 'end_col': 2, 'columns': ['A', 'B', 'C']}], 'ambiguities': []}
        fp = {'total_rows': 10, 'total_cols': 3, 'header_candidates': [{'row_index': 0}], 'blank_rows': [], 'blank_cols': []}

        report = self.gates.validate(plan, fp)
        assert isinstance(report, ValidationReport)
        assert report.all_passed or not report.requires_human_review

    def test_low_confidence_fails(self):
        plan = {'confidence': 0.2, 'subtables': [], 'ambiguities': []}
        fp = {'total_rows': 10, 'total_cols': 3, 'header_candidates': [], 'blank_rows': [], 'blank_cols': []}

        report = self.gates.validate(plan, fp)
        assert not report.all_passed
        assert report.requires_human_review

    def test_empty_subtables(self):
        plan = {'confidence': 0.7, 'subtables': [{'start_row': 0, 'end_row': 1, 'start_col': 0, 'end_col': 0, 'columns': ['A']}], 'ambiguities': []}
        fp = {'total_rows': 10, 'total_cols': 3, 'header_candidates': [{'row_index': 0}], 'blank_rows': [], 'blank_cols': []}

        report = self.gates.validate(plan, fp)
        gate_names = [g.gate_name for g in report.gates]
        assert 'subtable_structure' in gate_names

    def test_many_ambiguities_flagged(self):
        plan = {
            'confidence': 0.6,
            'subtables': [{'start_row': 0, 'end_row': 10, 'start_col': 0, 'end_col': 2, 'columns': ['A', 'B', 'C']}],
            'ambiguities': [f'issue_{i}' for i in range(50)],
        }
        fp = {'total_rows': 10, 'total_cols': 3, 'header_candidates': [{'row_index': 0}], 'blank_rows': [], 'blank_cols': []}

        report = self.gates.validate(plan, fp)
        unresolved_gate = next(g for g in report.gates if g.gate_name == 'unresolved_zones')
        assert not unresolved_gate.passed

    def test_to_dict(self):
        plan = {'confidence': 0.9, 'subtables': [{'start_row': 0, 'end_row': 5, 'start_col': 0, 'end_col': 2, 'columns': ['A', 'B', 'C']}], 'ambiguities': []}
        fp = {'total_rows': 10, 'total_cols': 3, 'header_candidates': [{'row_index': 0}], 'blank_rows': [], 'blank_cols': []}

        report = self.gates.validate(plan, fp)
        d = report.to_dict()
        assert isinstance(d, dict)
        assert 'gates' in d
        assert 'all_passed' in d


# ── LLM Service Tests ──

class TestLLMService:
    def test_parse_valid_json(self):
        service = LLMReconstructionService()
        response = '''
        {
            "subtables": [{"index": 0, "start_row": 0, "end_row": 5, "columns": ["A", "B"]}],
            "confidence": 0.85,
            "ambiguities": []
        }
        '''
        result = service._parse_llm_response(response)
        assert result['success'] is True
        assert result['confidence'] == 0.85
        assert len(result['subtables']) == 1

    def test_parse_json_in_code_block(self):
        service = LLMReconstructionService()
        response = '''Voici le resultat:
        ```json
        {"subtables": [], "confidence": 0.7}
        ```
        '''
        result = service._parse_llm_response(response)
        assert result['success'] is True
        assert result['confidence'] == 0.7

    def test_parse_invalid_json(self):
        service = LLMReconstructionService()
        result = service._parse_llm_response("Not JSON at all")
        assert result['success'] is False
        assert 'error' in result

    def test_format_correction_examples_empty(self):
        service = LLMReconstructionService()
        text = service._format_correction_examples([])
        assert "Aucun exemple" in text

    def test_format_correction_examples_with_data(self):
        service = LLMReconstructionService()
        examples = [{'correction_type': 'structural', 'description': 'Fix header', 'structural_before': {}, 'structural_after': {}}]
        text = service._format_correction_examples(examples)
        assert "Exemple 1" in text
        assert "Fix header" in text

    def test_build_structural_sample(self):
        service = LLMReconstructionService()
        fp = {
            'total_rows': 100,
            'total_cols': 5,
            'subtables': [{'start_row': 0, 'end_row': 50, 'start_col': 0, 'end_col': 4}],
            'header_candidates': [{'row_index': 0, 'score': 0.9, 'fill_ratio': 0.8}],
            'column_types': {'Nom': 'text', 'Montant': 'numeric'},
            'issues': ['2 cellules fusionnees'],
        }
        text = service._build_structural_sample(fp)
        assert '100 lignes' in text
        assert '5 colonnes' in text
        assert 'Sous-tableaux detectes' in text


# ── Correction Memory Tests ──

class TestCorrectionMemory:
    def test_fingerprint_to_text(self):
        memory = CorrectionMemory()
        fp = {
            'total_rows': 50,
            'total_cols': 4,
            'merged_cells': [{'start_row': 1, 'start_col': 1, 'end_row': 2, 'end_col': 3}],
            'blank_rows': [5],
            'blank_cols': [],
            'subtables': [{'start_row': 0, 'end_row': 4, 'start_col': 0, 'end_col': 3}],
            'header_candidates': [{'row_index': 0}],
            'column_types': {'A': 'text', 'B': 'numeric'},
            'issues': ['merged cells'],
        }
        text = memory._fingerprint_to_text(fp)
        assert 'rows=50' in text
        assert 'cols=4' in text
        assert 'merged=1' in text
        assert 'blank_rows=1' in text
        assert 'subtables=1' in text

    def test_hash_embedding_deterministic(self):
        memory = CorrectionMemory()
        text = "rows=50 | cols=4 | merged=1"
        e1 = memory._hash_embedding(text)
        e2 = memory._hash_embedding(text)
        assert e1 == e2
        assert len(e1) == 384

    def test_hash_embedding_normalized(self):
        memory = CorrectionMemory()
        e = memory._hash_embedding("test text")
        norm = sum(x**2 for x in e) ** 0.5
        assert abs(norm - 1.0) < 0.01

    def test_list_to_pgvector_sql(self):
        memory = CorrectionMemory()
        vec = [0.1, 0.2, 0.3]
        sql = memory._list_to_pgvector_sql(vec)
        assert sql.startswith('[')
        assert sql.endswith(']')
        assert '0.100000' in sql


# ── Integration Test ──

class TestStructureDetectionIntegration:
    def test_full_heuristic_flow(self, tmp_path):
        data = [
            ["Nom", "Prenom", "Montant"],
            ["Alice", "Bob", 100],
            ["Charlie", "Delta", 200],
            ["Eve", "Frank", 300],
        ]
        wb = Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        path = tmp_path / "integration_test.xlsx"
        wb.save(path)

        detector = HeuristicDetector()
        fp = detector.detect_from_file(str(path))

        assert fp.total_rows == 4
        assert fp.total_cols == 3
        assert fp.confidence > 0.5
        assert len(fp.header_candidates) > 0

        gates = ValidationGates()
        plan = {
            'confidence': fp.confidence,
            'subtables': [{'start_row': 0, 'end_row': 3, 'start_col': 0, 'end_col': 2, 'columns': ['Nom', 'Prenom', 'Montant']}],
            'ambiguities': [],
        }
        fp_dict = fp.to_dict()
        report = gates.validate(plan, fp_dict)
        assert report.all_passed or not report.requires_human_review


# ── Plan-to-Rules Tests ──

@pytest.mark.django_db
class TestPlanToRulesService:
    def setup_method(self):
        from django.contrib.auth.models import User
        self.user, _ = User.objects.get_or_create(
            username='test_plan_user',
            defaults={'email': 'test@example.com', 'is_staff': True},
        )
        from apps.ingestion.models import DataSource
        self.source, _ = DataSource.objects.get_or_create(
            name='Test Plan Source',
            defaults={
                'source_type': 'excel',
                'file_path': '/tmp/test.xlsx',
                'uploaded_by': self.user,
            },
        )

    def test_rename_columns_rule(self):
        from apps.nettoyage.structure_detection.plan_to_rules import PlanToRulesService
        service = PlanToRulesService()
        plan = {
            'subtables': [],
            'column_renames': {'Ancien Nom': 'Nouveau Nom', 'Prix HT': 'PrixHT'},
            'header_adjustments': [],
            'ambiguities': [],
        }
        result = service.apply_plan_to_source(plan=plan, source=self.source, user=self.user)
        assert result['total_rules'] >= 1
        rename_rule = next((r for r in result['rules_created'] if r['rule_type'] == 'rename_columns'), None)
        assert rename_rule is not None
        assert rename_rule['parameters']['mapping'] == {'Ancien Nom': 'Nouveau Nom', 'Prix HT': 'PrixHT'}

    def test_subtable_extraction_rule(self):
        from apps.nettoyage.structure_detection.plan_to_rules import PlanToRulesService
        service = PlanToRulesService()
        plan = {
            'subtables': [
                {'index': 0, 'start_row': 0, 'end_row': 5, 'start_col': 0, 'end_col': 3, 'column_names': ['A', 'B', 'C', 'D']},
                {'index': 1, 'start_row': 10, 'end_row': 15, 'start_col': 0, 'end_col': 2, 'column_names': ['X', 'Y', 'Z']},
            ],
            'column_renames': {},
            'header_adjustments': [],
            'ambiguities': [],
        }
        result = service.apply_plan_to_source(plan=plan, source=self.source, user=self.user)
        assert result['total_rules'] >= 1
        subtable_rule = next((r for r in result['rules_created'] if 'sous-table' in r['name']), None)
        assert subtable_rule is not None

    def test_empty_rows_rule_always_created(self):
        from apps.nettoyage.structure_detection.plan_to_rules import PlanToRulesService
        service = PlanToRulesService()
        plan = {
            'subtables': [],
            'column_renames': {},
            'header_adjustments': [],
            'ambiguities': [],
        }
        result = service.apply_plan_to_source(plan=plan, source=self.source, user=self.user)
        assert result['total_rules'] >= 1
        empty_rule = next((r for r in result['rules_created'] if r['rule_type'] == 'remove_empty_rows'), None)
        assert empty_rule is not None

    def test_pipeline_created(self):
        from apps.nettoyage.structure_detection.plan_to_rules import PlanToRulesService
        service = PlanToRulesService()
        plan = {
            'subtables': [],
            'column_renames': {'col1': 'col1_renamed'},
            'header_adjustments': [],
            'ambiguities': [],
        }
        result = service.apply_plan_to_source(plan=plan, source=self.source, user=self.user)
        assert result['pipeline_id'] is not None
        assert result['pipeline_name'] is not None
        from apps.nettoyage.models import CleaningPipeline
        pipeline = CleaningPipeline.objects.get(id=result['pipeline_id'])
        assert pipeline.rules.count() >= 1

    def test_execute_plan_returns_cleaning_result(self):
        from apps.nettoyage.structure_detection.plan_to_rules import PlanToRulesService
        from apps.nettoyage.services import CleaningError
        service = PlanToRulesService()
        plan = {
            'subtables': [],
            'column_renames': {},
            'header_adjustments': [],
            'ambiguities': [],
        }
        # Source has no raw rows, so execute_plan should raise CleaningError
        try:
            result = service.execute_plan(plan=plan, source=self.source, user=self.user)
            # If it doesn't raise, it should return no_rules status
            assert result['status'] == 'no_rules'
        except CleaningError:
            # Expected: source has no raw data rows
            pass


# ---- Pivot Detection Tests ----

class TestPivotDetection:
    def setup_method(self):
        self.detector = HeuristicDetector()

    def _create_pivot_excel(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Produit", "Janvier", "Fevrier", "Mars"])
        ws.append(["A", 100, "", ""])
        ws.append(["B", "", 200, ""])
        ws.append(["C", "", "", 300])
        ws.append(["D", 50, "", ""])
        ws.append(["E", "", 150, ""])
        path = tmp_path / "test_pivot.xlsx"
        wb.save(path)
        return str(path)

    def _create_hierarchical_pivot_excel(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["", "Janvier", "Janvier", "Fevrier", "Fevrier"])
        ws.append(["Produit", "Nord", "Sud", "Nord", "Sud"])
        ws.append(["A", 10, 20, 30, 40])
        ws.append(["B", 15, 25, 35, 45])
        ws.append(["C", 5, 10, 15, 20])
        path = tmp_path / "test_hier_pivot.xlsx"
        wb.save(path)
        return str(path)

    def test_detect_sparse_pivot(self, tmp_path):
        path = self._create_pivot_excel(tmp_path)
        fp = self.detector.detect_from_file(path)
        assert len(fp.sparse_pivot_candidates) >= 1
        assert fp.force_llm is True

    def test_pivot_candidate_has_correct_columns(self, tmp_path):
        path = self._create_pivot_excel(tmp_path)
        fp = self.detector.detect_from_file(path)
        candidate = fp.sparse_pivot_candidates[0]
        assert 'Produit' in candidate['dimension_headers']
        assert any(h in candidate['value_headers'] for h in ['Janvier', 'Fevrier', 'Mars'])

    def test_pivot_sparsity_score(self, tmp_path):
        path = self._create_pivot_excel(tmp_path)
        fp = self.detector.detect_from_file(path)
        candidate = fp.sparse_pivot_candidates[0]
        assert candidate['single_value_ratio'] >= 0.9

    def test_hierarchical_headers_detected(self, tmp_path):
        path = self._create_hierarchical_pivot_excel(tmp_path)
        fp = self.detector.detect_from_file(path)
        assert len(fp.hierarchical_headers) >= 1

    def test_no_pivot_in_dense_data(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Nom", "Valeur1", "Valeur2", "Valeur3"])
        ws.append(["A", 100, 200, 300])
        ws.append(["B", 150, 250, 350])
        ws.append(["C", 110, 210, 310])
        path = tmp_path / "test_dense.xlsx"
        wb.save(path)
        fp = self.detector.detect_from_file(str(path))
        assert len(fp.sparse_pivot_candidates) == 0

    def test_detect_from_dataframe_pivot(self):
        df = pd.DataFrame({
            'Produit': ['A', 'B', 'C'],
            'Janvier': [100, None, None],
            'Fevrier': [None, 200, None],
            'Mars': [None, None, 300],
        })
        fp = self.detector.detect_from_dataframe(df)
        assert len(fp.sparse_pivot_candidates) >= 1

    def test_fingerprint_to_dict_includes_pivot(self, tmp_path):
        path = self._create_pivot_excel(tmp_path)
        fp = self.detector.detect_from_file(path)
        d = fp.to_dict()
        assert 'sparse_pivot_candidates' in d
        assert 'hierarchical_headers' in d
        assert 'force_llm' in d


# ---- Pivot Transformer Tests ----

class TestPivotTransformer:
    def setup_method(self):
        self.transformer = PivotTransformer()

    def test_unpivot_basic(self):
        df = pd.DataFrame({
            'Produit': ['A', 'B', 'C'],
            'Janvier': [100, 0, 0],
            'Fevrier': [0, 200, 0],
            'Mars': [0, 0, 300],
        })
        result = self.transformer.unpivot(
            df,
            id_vars=['Produit'],
            value_vars=['Janvier', 'Fevrier', 'Mars'],
            var_name='Mois',
            value_name='Ventes',
        )
        assert result is not None
        assert len(result) > 0
        assert 'Mois' in result.columns
        assert 'Ventes' in result.columns

    def test_unpivot_drops_na(self):
        df = pd.DataFrame({
            'Produit': ['A', 'B', 'C'],
            'Janvier': [100, None, None],
            'Fevrier': [None, 200, None],
            'Mars': [None, None, 300],
        })
        result = self.transformer.unpivot(
            df,
            id_vars=['Produit'],
            value_vars=['Janvier', 'Fevrier', 'Mars'],
            var_name='Mois',
            value_name='Ventes',
        )
        assert result is not None
        assert all(result['Ventes'].notna())

    def test_unpivot_from_mapping(self):
        df = pd.DataFrame({
            'Product': ['X', 'Y'],
            'Q1': [10, 20],
            'Q2': [30, 40],
        })
        mapping = {
            'colonnes_identifiantes': ['Product'],
            'colonnes_valeurs': ['Q1', 'Q2'],
            'nom_nouvelle_colonne_dimension': 'Trimestre',
            'nom_nouvelle_colonne_valeur': 'Ventes',
        }
        result = self.transformer.unpivot_from_mapping(df, mapping)
        assert result is not None
        assert 'Trimestre' in result.columns
        assert 'Ventes' in result.columns
        assert len(result) == 4

    def test_unpivot_empty_df(self):
        df = pd.DataFrame()
        result = self.transformer.unpivot(df, id_vars=[], value_vars=[])
        assert result is None

    def test_unpivot_missing_columns(self):
        df = pd.DataFrame({'A': [1], 'B': [2]})
        result = self.transformer.unpivot(df, id_vars=['X'], value_vars=['Y'])
        assert result is None

    def test_validate_unpivot_result(self):
        original_df = pd.DataFrame({
            'Produit': ['A', 'B', 'C'],
            'Janvier': [100, 0, 0],
            'Fevrier': [0, 200, 0],
            'Mars': [0, 0, 300],
        })
        unpivoted_df = self.transformer.unpivot(
            original_df,
            id_vars=['Produit'],
            value_vars=['Janvier', 'Fevrier', 'Mars'],
        )
        mapping = {
            'colonnes_identifiantes': ['Produit'],
            'colonnes_valeurs': ['Janvier', 'Fevrier', 'Mars'],
            'nom_nouvelle_colonne_dimension': 'Mois',
            'nom_nouvelle_colonne_valeur': 'Ventes',
        }
        validation = self.transformer.validate_unpivot_result(original_df, unpivoted_df, mapping)
        assert validation['valid'] is True
        assert validation['row_count_ok'] is True

    def test_validate_unpivot_row_count_mismatch(self):
        original_df = pd.DataFrame({
            'X': ['A', 'B'],
            'Y': [10, None],
            'Z': [None, 20],
        })
        unpivoted_df = pd.DataFrame({
            'X': ['A', 'B', 'C'],
            'Dimension': ['Y', 'Z', 'Y'],
            'Valeur': [10, 20, 30],
        })
        mapping = {
            'colonnes_identifiantes': ['X'],
            'colonnes_valeurs': ['Y', 'Z'],
        }
        validation = self.transformer.validate_unpivot_result(original_df, unpivoted_df, mapping)
        assert validation['row_count_ok'] is False


# ---- Unpivot Validation Gate Tests ----

class TestUnpivotValidationGate:
    def setup_method(self):
        self.gates = ValidationGates()

    def test_unpivot_valid_mapping_passes(self):
        plan = {
            'confidence': 0.85,
            'subtables': [{'start_row': 0, 'end_row': 5, 'start_col': 0, 'end_col': 3, 'columns': ['A', 'B', 'C', 'D']}],
            'ambiguities': [],
            'type_transformation': 'unpivot',
            'mapping_pivot': {
                'colonnes_identifiantes': ['A'],
                'colonnes_valeurs': ['B', 'C', 'D'],
                'nom_nouvelle_colonne_dimension': 'Categorie',
                'nom_nouvelle_colonne_valeur': 'Montant',
            },
        }
        fp = {
            'total_rows': 10,
            'total_cols': 4,
            'header_candidates': [{'row_index': 0}],
            'blank_rows': [],
            'blank_cols': [],
            'column_types': {'A': 'text', 'B': 'numeric', 'C': 'numeric', 'D': 'numeric'},
        }
        report = self.gates.validate(plan, fp)
        gate_names = [g.gate_name for g in report.gates]
        assert 'unpivot_mapping' in gate_names
        assert 'unpivot_column_existence' in gate_names

    def test_unpivot_missing_columns_fails(self):
        plan = {
            'confidence': 0.85,
            'subtables': [{'start_row': 0, 'end_row': 5, 'start_col': 0, 'end_col': 3, 'columns': ['A', 'B', 'C', 'D']}],
            'ambiguities': [],
            'type_transformation': 'unpivot',
            'mapping_pivot': {
                'colonnes_identifiantes': ['X'],
                'colonnes_valeurs': ['Y', 'Z'],
                'nom_nouvelle_colonne_dimension': 'Dim',
                'nom_nouvelle_colonne_valeur': 'Val',
            },
        }
        fp = {
            'total_rows': 10,
            'total_cols': 4,
            'header_candidates': [{'row_index': 0}],
            'blank_rows': [],
            'blank_cols': [],
            'column_types': {'A': 'text', 'B': 'numeric', 'C': 'numeric', 'D': 'numeric'},
        }
        report = self.gates.validate(plan, fp)
        col_exist_gate = next(g for g in report.gates if g.gate_name == 'unpivot_column_existence')
        assert not col_exist_gate.passed

    def test_unpivot_empty_mapping_fails(self):
        plan = {
            'confidence': 0.85,
            'subtables': [{'start_row': 0, 'end_row': 5, 'start_col': 0, 'end_col': 3, 'columns': ['A', 'B', 'C', 'D']}],
            'ambiguities': [],
            'type_transformation': 'unpivot',
            'mapping_pivot': {
                'colonnes_identifiantes': [],
                'colonnes_valeurs': [],
                'nom_nouvelle_colonne_dimension': '',
                'nom_nouvelle_colonne_valeur': '',
            },
        }
        fp = {
            'total_rows': 10,
            'total_cols': 4,
            'header_candidates': [{'row_index': 0}],
            'blank_rows': [],
            'blank_cols': [],
            'column_types': {'A': 'text', 'B': 'numeric'},
        }
        report = self.gates.validate(plan, fp)
        mapping_gate = next(g for g in report.gates if g.gate_name == 'unpivot_mapping')
        assert not mapping_gate.passed

    def test_no_unpivot_skips_gate(self):
        plan = {
            'confidence': 0.9,
            'subtables': [{'start_row': 0, 'end_row': 5, 'start_col': 0, 'end_col': 2, 'columns': ['A', 'B', 'C']}],
            'ambiguities': [],
        }
        fp = {
            'total_rows': 10,
            'total_cols': 3,
            'header_candidates': [{'row_index': 0}],
            'blank_rows': [],
            'blank_cols': [],
            'column_types': {'A': 'text', 'B': 'numeric'},
        }
        report = self.gates.validate(plan, fp)
        gate_names = [g.gate_name for g in report.gates]
        assert 'unpivot_mapping' not in gate_names
        assert 'unpivot_column_existence' not in gate_names


# ---- LLM Response Parsing with Pivot ----

class TestLLMPivotParsing:
    def test_parse_unpivot_response(self):
        service = LLMReconstructionService()
        response = '''
        {
            "subtables": [{"start_row": 0, "end_row": 5, "start_col": 0, "end_col": 5, "columns": ["Produit", "Janvier", "Fevrier", "Mars"]}],
            "confidence": 0.92,
            "type_transformation": "unpivot",
            "mapping_pivot": {
                "colonnes_identifiantes": ["Produit"],
                "colonnes_valeurs": ["Janvier", "Fevrier", "Mars"],
                "nom_nouvelle_colonne_dimension": "Mois",
                "nom_nouvelle_colonne_valeur": "Ventes"
            },
            "ambiguities": []
        }
        '''
        result = service._parse_llm_response(response)
        assert result['success'] is True
        assert result['type_transformation'] == 'unpivot'
        assert result['mapping_pivot'] is not None
        assert result['mapping_pivot']['colonnes_identifiantes'] == ['Produit']
        assert result['mapping_pivot']['colonnes_valeurs'] == ['Janvier', 'Fevrier', 'Mars']

    def test_parse_flat_response_no_pivot(self):
        service = LLMReconstructionService()
        response = '''
        {
            "subtables": [{"start_row": 0, "end_row": 5, "columns": ["A", "B", "C"]}],
            "confidence": 0.88,
            "ambiguities": []
        }
        '''
        result = service._parse_llm_response(response)
        assert result['success'] is True
        assert result['type_transformation'] is None
        assert result['mapping_pivot'] is None

    def test_structural_sample_includes_pivot(self):
        service = LLMReconstructionService()
        fp = {
            'total_rows': 10,
            'total_cols': 5,
            'subtables': [],
            'header_candidates': [{'row_index': 0, 'score': 0.8, 'fill_ratio': 0.7}],
            'column_types': {},
            'issues': [],
            'sparse_pivot_candidates': [
                {'group_headers': ['Janvier', 'Fevrier', 'Mars'], 'sparsity_score': 0.92}
            ],
            'hierarchical_headers': [{'row_index': 0, 'values': ['', 'Janvier', 'Fevrier', 'Mars'], 'non_empty_count': 3}],
        }
        text = service._build_structural_sample(fp)
        assert 'Pivot epars detecte' in text
        assert 'En-tetes hierarchiques' in text


# ---- Cell-Level Transformer Tests ----

class TestExtractLabeledFields:
    def test_basic_extraction(self):
        texte = "Name Hussein Hakeem Address 123 Main St Age 17 Gender Male"
        labels = ["Name", "Address", "Age", "Gender"]
        result = extract_labeled_fields(texte, labels)
        assert result["Name"] == "Hussein Hakeem"
        assert result["Address"] == "123 Main St"
        assert result["Age"] == "17"
        assert result["Gender"] == "Male"

    def test_missing_label(self):
        texte = "Name John Age 30"
        labels = ["Name", "Address", "Age"]
        result = extract_labeled_fields(texte, labels)
        assert result["Name"] == "John"
        assert result["Address"] == ""
        assert result["Age"] == "30"

    def test_empty_text(self):
        result = extract_labeled_fields("", ["Name", "Age"])
        assert result == {"Name": "", "Age": ""}

    def test_label_at_end(self):
        texte = "Name Alice Age"
        result = extract_labeled_fields(texte, ["Name", "Age"])
        assert result["Name"] == "Alice"
        assert result["Age"] == ""

    def test_case_insensitive(self):
        texte = "name Bob age 25"
        result = extract_labeled_fields(texte, ["Name", "Age"])
        assert result["Name"] == "Bob"
        assert result["Age"] == "25"


class TestFixAmbiguousChars:
    def test_obvious_correction(self):
        result = fix_ambiguous_numeric_chars("41oo", {"o": "0", "O": "0"})
        assert result["corrected"] == "4100"
        assert result["was_corrected"] is True
        assert result["needs_review"] is False

    def test_uncertain_value(self):
        result = fix_ambiguous_numeric_chars(
            "0.S", cas_incertains=["0.S"]
        )
        assert result["corrected"] == "0.S"
        assert result["was_corrected"] is False
        assert result["needs_review"] is True

    def test_no_ambiguous_chars(self):
        result = fix_ambiguous_numeric_chars("12345")
        assert result["corrected"] == "12345"
        assert result["was_corrected"] is False

    def test_default_substitutions(self):
        result = fix_ambiguous_numeric_chars("iO")
        assert result["corrected"] == "10"
        assert result["was_corrected"] is True

    def test_mixed_ambiguous_and_uncertain(self):
        result = fix_ambiguous_numeric_chars(
            "iO", cas_incertains=["iO"]
        )
        assert result["needs_review"] is True
        assert result["corrected"] == "iO"


class TestSplitValueUnit:
    def test_basic_split(self):
        result = split_value_unit("0Bottle")
        assert result["nombre"] == "0"
        assert result["texte"] == "Bottle"

    def test_number_with_decimal(self):
        result = split_value_unit("3.5kg")
        assert result["nombre"] == "3.5"
        assert result["texte"] == "kg"

    def test_no_match(self):
        result = split_value_unit("Hello World")
        assert result is None

    def test_text_only(self):
        result = split_value_unit("Bottle")
        assert result is None

    def test_comma_decimal(self):
        result = split_value_unit("2,5L")
        assert result["nombre"] == "2.5"
        assert result["texte"] == "L"

    def test_empty_string(self):
        result = split_value_unit("")
        assert result is None


class TestExplodeDelimitedLists:
    def test_basic_explode(self):
        ligne = {
            "Category": "Binders | Art | Phones",
            "Amount": "609.98 | 5.48 | 391.98",
            "Order ID": "ORD-001",
        }
        result = explode_delimited_lists(
            ligne,
            colonnes_liees=["Category", "Amount"],
            delimiteur="|",
            colonnes_a_repeter=["Order ID"],
        )
        assert len(result) == 3
        assert result[0]["Category"] == "Binders"
        assert result[0]["Amount"] == "609.98"
        assert result[0]["Order ID"] == "ORD-001"
        assert result[2]["Category"] == "Phones"

    def test_length_mismatch_raises(self):
        ligne = {
            "A": "x | y",
            "B": "1 | 2 | 3",
        }
        with pytest.raises(ValueError, match="Longueurs de listes incoherentes"):
            explode_delimited_lists(
                ligne,
                colonnes_liees=["A", "B"],
                delimiteur="|",
                colonnes_a_repeter=[],
            )

    def test_single_element(self):
        ligne = {"Category": "Only One", "Amount": "99.99"}
        result = explode_delimited_lists(
            ligne,
            colonnes_liees=["Category", "Amount"],
            delimiteur="|",
            colonnes_a_repeter=[],
        )
        assert len(result) == 1
        assert result[0]["Category"] == "Only One"

    def test_empty_value(self):
        ligne = {"Category": "", "Amount": ""}
        result = explode_delimited_lists(
            ligne,
            colonnes_liees=["Category", "Amount"],
            delimiteur="|",
            colonnes_a_repeter=[],
        )
        assert len(result) == 1
        assert result[0]["Category"] == ""


class TestCellTransformerEngine:
    def setup_method(self):
        self.engine = CellTransformerEngine()

    def test_apply_text_field_extraction(self):
        df = pd.DataFrame({
            "ID": [1, 2],
            "RawText": [
                "Name Alice Age 25",
                "Name Bob Age 30",
            ],
        })
        transformation = {
            "colonne_source": "RawText",
            "sous_type": "extraction_champs_texte_libre",
            "colonnes_resultantes": ["Name", "Age"],
            "details": {"labels_detectes": ["Name", "Age"]},
        }
        result = self.engine.apply_text_field_extraction(df, transformation)
        assert "Name" in result.columns
        assert "Age" in result.columns
        assert "RawText" not in result.columns
        assert result.iloc[0]["Name"] == "Alice"
        assert result.iloc[0]["Age"] == "25"

    def test_apply_char_correction(self):
        df = pd.DataFrame({
            "Value": ["41oo", "1234", "iO"],
        })
        transformation = {
            "colonne_source": "Value",
            "sous_type": "correction_caracteres_ambigus",
            "colonnes_resultantes": ["Value"],
            "details": {
                "substitutions_types": {"o": "0", "O": "0", "i": "1"},
                "cas_incertains": ["iO"],
            },
        }
        result, review = self.engine.apply_char_correction(df, transformation)
        assert result.iloc[0]["Value"] == "4100"
        assert result.iloc[1]["Value"] == "1234"
        assert "iO" in review

    def test_apply_value_unit_split(self):
        df = pd.DataFrame({
            "Combined": ["0Bottle", "5kg", "12items"],
        })
        transformation = {
            "colonne_source": "Combined",
            "sous_type": "scission_valeur_unite",
            "colonnes_resultantes": ["Quantity", "Measure"],
            "details": {
                "colonne_nombre": "Quantity",
                "colonne_texte": "Measure",
            },
        }
        result = self.engine.apply_value_unit_split(df, transformation)
        assert "Quantity" in result.columns
        assert "Measure" in result.columns
        assert "Combined" not in result.columns
        assert result.iloc[0]["Quantity"] == "0"
        assert result.iloc[0]["Measure"] == "Bottle"

    def test_apply_delimited_list_explode(self):
        df = pd.DataFrame({
            "Category": ["A | B | C", "X | Y"],
            "Amount": ["10 | 20 | 30", "100 | 200"],
            "Order ID": ["ORD-1", "ORD-2"],
        })
        transformation = {
            "colonne_source": "Category",
            "sous_type": "explosion_liste_delimitee",
            "colonnes_resultantes": ["Category", "Amount"],
            "details": {
                "colonnes_liees": ["Category", "Amount"],
                "delimiteur_detecte": "|",
                "colonnes_a_repeter": ["Order ID"],
            },
        }
        result, mismatches = self.engine.apply_delimited_list_explode(df, transformation)
        assert len(result) == 5
        assert len(mismatches) == 0
        assert result.iloc[0]["Order ID"] == "ORD-1"
        assert result.iloc[3]["Order ID"] == "ORD-2"


# ---- Cell-Level Validation Gate Tests ----

class TestCellLevelValidationGates:
    def setup_method(self):
        self.gates = ValidationGates()

    def test_text_extraction_gate_passes(self):
        plan = {
            'confidence': 0.85,
            'subtables': [{'start_row': 0, 'end_row': 5, 'start_col': 0, 'end_col': 2, 'columns': ['A', 'B', 'C']}],
            'ambiguities': [],
            'transformations_cellule': [{
                'colonne_source': 'RawText',
                'sous_type': 'extraction_champs_texte_libre',
                'colonnes_resultantes': ['Name', 'Address', 'Age'],
                'details': {'labels_detectes': ['Name', 'Address', 'Age']},
                'niveau_confiance': 0.9,
                'echantillon_ambigu': [],
            }],
        }
        fp = {'total_rows': 10, 'total_cols': 3, 'header_candidates': [{'row_index': 0}], 'blank_rows': [], 'blank_cols': [], 'column_types': {'Name': 'text', 'Address': 'text', 'Age': 'text'}}
        report = self.gates.validate(plan, fp)
        gate_names = [g.gate_name for g in report.gates]
        assert 'text_extraction_empty_rate' in gate_names

    def test_char_correction_gate_fails_low_confidence(self):
        plan = {
            'confidence': 0.85,
            'subtables': [{'start_row': 0, 'end_row': 5, 'start_col': 0, 'end_col': 2, 'columns': ['A', 'B', 'C']}],
            'ambiguities': [],
            'transformations_cellule': [{
                'colonne_source': 'Value',
                'sous_type': 'correction_caracteres_ambigus',
                'colonnes_resultantes': ['Value'],
                'details': {'substitutions_types': {'o': '0'}, 'cas_incertains': []},
                'niveau_confiance': 0.3,
                'echantillon_ambigu': [],
            }],
        }
        fp = {'total_rows': 10, 'total_cols': 3, 'header_candidates': [{'row_index': 0}], 'blank_rows': [], 'blank_cols': [], 'column_types': {'Value': 'text'}}
        report = self.gates.validate(plan, fp)
        gate = next(g for g in report.gates if g.gate_name == 'char_correction_completeness')
        assert not gate.passed

    def test_char_correction_gate_passes(self):
        plan = {
            'confidence': 0.85,
            'subtables': [{'start_row': 0, 'end_row': 5, 'start_col': 0, 'end_col': 2, 'columns': ['A', 'B', 'C']}],
            'ambiguities': [],
            'transformations_cellule': [{
                'colonne_source': 'Value',
                'sous_type': 'correction_caracteres_ambigus',
                'colonnes_resultantes': ['Value'],
                'details': {'substitutions_types': {'o': '0'}, 'cas_incertains': []},
                'niveau_confiance': 0.95,
                'echantillon_ambigu': [],
            }],
        }
        fp = {'total_rows': 10, 'total_cols': 3, 'header_candidates': [{'row_index': 0}], 'blank_rows': [], 'blank_cols': [], 'column_types': {'Value': 'text'}}
        report = self.gates.validate(plan, fp)
        gate = next(g for g in report.gates if g.gate_name == 'char_correction_completeness')
        assert gate.passed

    def test_value_unit_gate_passes(self):
        plan = {
            'confidence': 0.85,
            'subtables': [{'start_row': 0, 'end_row': 5, 'start_col': 0, 'end_col': 2, 'columns': ['A', 'B', 'C']}],
            'ambiguities': [],
            'transformations_cellule': [{
                'colonne_source': 'Combined',
                'sous_type': 'scission_valeur_unite',
                'colonnes_resultantes': ['Quantity', 'Measure'],
                'details': {'colonne_nombre': 'Quantity', 'colonne_texte': 'Measure'},
                'niveau_confiance': 0.9,
                'echantillon_ambigu': [],
            }],
        }
        fp = {'total_rows': 10, 'total_cols': 3, 'header_candidates': [{'row_index': 0}], 'blank_rows': [], 'blank_cols': [], 'column_types': {'Quantity': 'numeric', 'Measure': 'text'}}
        report = self.gates.validate(plan, fp)
        gate = next(g for g in report.gates if g.gate_name == 'value_unit_numeric')
        assert gate.passed

    def test_delimited_list_gate_passes(self):
        plan = {
            'confidence': 0.85,
            'subtables': [{'start_row': 0, 'end_row': 5, 'start_col': 0, 'end_col': 2, 'columns': ['A', 'B', 'C']}],
            'ambiguities': [],
            'transformations_cellule': [{
                'colonne_source': 'Category',
                'sous_type': 'explosion_liste_delimitee',
                'colonnes_resultantes': ['Category', 'Amount'],
                'details': {
                    'colonnes_liees': ['Category', 'Amount'],
                    'delimiteur_detecte': '|',
                    'colonnes_a_repeter': ['Order ID'],
                },
                'niveau_confiance': 0.85,
                'echantillon_ambigu': [],
            }],
        }
        fp = {'total_rows': 10, 'total_cols': 3, 'header_candidates': [{'row_index': 0}], 'blank_rows': [], 'blank_cols': [], 'column_types': {'Category': 'text', 'Amount': 'numeric'}}
        report = self.gates.validate(plan, fp)
        gate = next(g for g in report.gates if g.gate_name == 'delimited_list_integrity')
        assert gate.passed

    def test_delimited_list_gate_fails_too_few_cols(self):
        plan = {
            'confidence': 0.85,
            'subtables': [{'start_row': 0, 'end_row': 5, 'start_col': 0, 'end_col': 2, 'columns': ['A', 'B', 'C']}],
            'ambiguities': [],
            'transformations_cellule': [{
                'colonne_source': 'Category',
                'sous_type': 'explosion_liste_delimitee',
                'colonnes_resultantes': ['Category'],
                'details': {
                    'colonnes_liees': ['Category'],
                    'delimiteur_detecte': '|',
                    'colonnes_a_repeter': [],
                },
                'niveau_confiance': 0.85,
                'echantillon_ambigu': [],
            }],
        }
        fp = {'total_rows': 10, 'total_cols': 1, 'header_candidates': [{'row_index': 0}], 'blank_rows': [], 'blank_cols': [], 'column_types': {'Category': 'text'}}
        report = self.gates.validate(plan, fp)
        gate = next(g for g in report.gates if g.gate_name == 'delimited_list_integrity')
        assert not gate.passed


# ---- LLM Response Parsing with Cell-Level ----

class TestLLMCellLevelParsing:
    def test_parse_cell_level_response(self):
        service = LLMReconstructionService()
        response = '''
        {
            "subtables": [{"start_row": 0, "end_row": 5, "columns": ["ID", "RawText"]}],
            "confidence": 0.88,
            "type_transformation": "cell_level",
            "transformations_cellule": [
                {
                    "colonne_source": "RawText",
                    "sous_type": "extraction_champs_texte_libre",
                    "colonnes_resultantes": ["Name", "Address", "Age"],
                    "details": {
                        "labels_detectes": ["Name", "Address", "Age"],
                        "logique_decoupage": "chaque label marque le debut d'un champ"
                    },
                    "niveau_confiance": 0.92,
                    "echantillon_ambigu": ["Name incomplete entry"]
                }
            ],
            "ambiguities": []
        }
        '''
        result = service._parse_llm_response(response)
        assert result['success'] is True
        assert result['type_transformation'] == 'cell_level'
        assert result['transformations_cellule'] is not None
        assert len(result['transformations_cellule']) == 1
        ct = result['transformations_cellule'][0]
        assert ct['sous_type'] == 'extraction_champs_texte_libre'
        assert ct['colonnes_resultantes'] == ['Name', 'Address', 'Age']

    def test_parse_mixed_response(self):
        service = LLMReconstructionService()
        response = '''
        {
            "subtables": [{"start_row": 0, "end_row": 10, "columns": ["A", "B", "C"]}],
            "confidence": 0.9,
            "type_transformation": "mixed",
            "mapping_pivot": {
                "colonnes_identifiantes": ["A"],
                "colonnes_valeurs": ["B", "C"],
                "nom_nouvelle_colonne_dimension": "Dim",
                "nom_nouvelle_colonne_valeur": "Val"
            },
            "transformations_cellule": [
                {
                    "colonne_source": "A",
                    "sous_type": "correction_caracteres_ambigus",
                    "colonnes_resultantes": ["A"],
                    "details": {
                        "substitutions_types": {"o": "0"},
                        "cas_incertains": ["0.S"]
                    },
                    "niveau_confiance": 0.85,
                    "echantillon_ambigu": ["0.S"]
                }
            ],
            "ambiguities": []
        }
        '''
        result = service._parse_llm_response(response)
        assert result['success'] is True
        assert result['type_transformation'] == 'mixed'
        assert result['mapping_pivot'] is not None
        assert result['transformations_cellule'] is not None
        assert result['transformations_cellule'][0]['sous_type'] == 'correction_caracteres_ambigus'
